import json
import sys
import openpyxl

xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "112211.xlsx"
out_path = sys.argv[2] if len(sys.argv) > 2 else "scripts/excel-prices-parsed.json"

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb.active
rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    cells = [None if c is None else str(c).strip() for c in row]
    if any(c for c in cells if c):
        rows.append(cells)

with open(out_path, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print(len(rows))
